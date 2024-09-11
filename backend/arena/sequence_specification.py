import os
import pandas as pd
import openpyxl as px

import sys

#sys.path.insert(1, "../../backend")
#from constants import RED, RESET


class SequenceSpecification:
    def __init__(self, filePath) -> None:
        self.name = os.path.basename(filePath)
        self.sequenceSheet = get_sequence_sheet(filePath)
        self.statements = {}
        
        self.statements = {}
        i = 0
        for row in self.sequenceSheet:
            statement = Statement(
                i,
                row[0],
                row[1],
                row[2],
                row[3:],
            )
            statement.inputParams = [param for param in statement.inputParams if param is not None]
            self.statements[i] = statement
            i += 1
        validSheet = False
        for statement in self.statements.values():
            if statement.methodName == "create" and "." not in statement.instanceParam:
                validSheet = True
        if not validSheet:
            raise ValueError("Invalid sequence sheet. No create statement found.")
                
    def resolve_reference(self, cell: str):
        """
        Resolve cell references like A1, B2, etc.
        """
        row = int(cell[1:]) - 1
        if row >= len(self.statements):
            # row does not exist
            return cell
        col = ord(cell[0].upper()) - 65
        if col >= len(self.statements[row].inputParams)+3:
            # column does not exist
            return cell
        if col == 0:
            if self.statements[row].output is not None:
                return self.statements[row].output
            else:
                return self.statements[row].oracleValue
        elif col == 1:
            return self.statements[row].methodName
        elif col == 2:
            return self.statements[row].instanceParam
        elif col >= 3:
            return self.statements[row].inputParams[col-3]
        else:
            return cell
        
    def printSequenceSheet(self):
        """
        Print the sequence sheet.
        """
        for statement in self.statements.values():
            print(statement)

    def reset(self):
        """
        Reset the sequence sheet, needed to remove output values and resolved references.
        """
        self.statements = {}
        i = 0
        for row in self.sequenceSheet:
            statement = Statement(
                i,
                row[0],
                row[1],
                row[2],
                row[3:],
            )
            statement.inputParams = [param for param in statement.inputParams if param is not None]
            self.statements[i] = statement
            i += 1
        validSheet = False
        for statement in self.statements.values():
            if statement.methodName == "create" and "." not in statement.instanceParam:
                validSheet = True
        if not validSheet:
            raise ValueError("Invalid sequence sheet. No create statement found.")

class Statement:
    def __init__(self, position, oracleValue, methodName, instanceParam, inputParams) -> None:
        self.position = position
        self.oracleValue = oracleValue
        self.methodName = methodName
        self.instanceParam = instanceParam
        self.inputParams = inputParams
        self.output = None
    
    def __repr__(self) -> str:
        return f"[{self.position}] Oracle Value: {self.oracleValue} | Output: {self.output} | Method Name: {self.methodName} | Instance Param: {self.instanceParam} | Input Params: {self.inputParams}"


def get_sequence_sheet(path) -> list:
    """
    Read in excel/csv file and return sequence sheet as a list.
    """
    _, file_extension = os.path.splitext(path)
    
    sheet_array = []
            
    if file_extension in [".xls", ".xlsx", ".xlsm", ".xlsb"]:
        # read in excel file
        wb = px.load_workbook(path)
        sheet = wb.active
        # convert to arrays in array
        for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row, values_only=True):
            row_array = []
            for cell in row:
                row_array.append(cell)
            sheet_array.append(row_array)
    elif file_extension in [".csv"]:
        # read in csv file in array of arrays
        with open(path, "r") as file:
            for line in file:
                row_array = line.strip().split(";")
                sheet_array.append(row_array)
    else:
        raise ValueError("Unsupported file type")    
    
    return sheet_array